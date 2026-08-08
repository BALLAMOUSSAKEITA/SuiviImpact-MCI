"""Mise en forme institutionnelle des exports Excel (openpyxl)."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

EXPORT_MINISTRY = (
    "Le ministère de l'Industrie et du Commerce, "
    "Bureau des stratégies de développement"
)

COLOR_GREEN = "009959"
COLOR_GREEN_DARK = "007A47"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_TITLE = "1A3D2E"
COLOR_MUTED = "5A6B63"
COLOR_ROW_ALT = "F4FAF7"
COLOR_BORDER = "D4E8DE"

_FONT = "Calibri"


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def prepare_branded_sheet(
    ws: Worksheet,
    *,
    report_title: str,
    headers: list[str],
    subtitle: str | None = None,
) -> int:
    """
    En-tête institutionnel + ligne de titres de colonnes.
    Retourne l'index de la première ligne de données.
    """
    ncols = max(len(headers), 1)
    last = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last}1")
    org = ws["A1"]
    org.value = EXPORT_MINISTRY
    org.font = Font(name=_FONT, size=11, italic=True, color=COLOR_MUTED)
    org.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last}2")
    bar = ws["A2"]
    bar.fill = PatternFill(
        start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid"
    )
    ws.row_dimensions[2].height = 6

    ws.merge_cells(f"A3:{last}3")
    title_cell = ws["A3"]
    title_cell.value = report_title
    title_cell.font = Font(name=_FONT, size=16, bold=True, color=COLOR_GREEN_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 28

    header_row = 5
    if subtitle:
        ws.merge_cells(f"A4:{last}4")
        sub = ws["A4"]
        sub.value = subtitle
        sub.font = Font(name=_FONT, size=11, color=COLOR_MUTED)
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 22
    else:
        header_row = 4

    header_fill = PatternFill(
        start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid"
    )
    header_font = Font(name=_FONT, bold=True, color=COLOR_HEADER_TEXT, size=11)
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _thin_border()

    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return header_row + 1


def write_data_rows(ws: Worksheet, start_row: int, rows: list[tuple]) -> None:
    if not rows:
        return
    ncols = max(len(r) for r in rows)
    alt_fill = PatternFill(
        start_color=COLOR_ROW_ALT, end_color=COLOR_ROW_ALT, fill_type="solid"
    )
    body_font = Font(name=_FONT, size=10, color=COLOR_TITLE)
    border = _thin_border()
    body_align = Alignment(vertical="top", wrap_text=True)

    for i, row in enumerate(rows):
        excel_row = start_row + i
        use_alt = i % 2 == 1
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border
            if use_alt:
                cell.fill = alt_fill

    autofit_columns(ws, ncols)


def write_row_cells(
    ws: Worksheet,
    row_idx: int,
    values: tuple,
    *,
    zebra_index: int,
) -> None:
    """Une ligne de données (exports écrits cellule par cellule)."""
    alt_fill = PatternFill(
        start_color=COLOR_ROW_ALT, end_color=COLOR_ROW_ALT, fill_type="solid"
    )
    body_font = Font(name=_FONT, size=10, color=COLOR_TITLE)
    border = _thin_border()
    body_align = Alignment(vertical="top", wrap_text=True)
    use_alt = zebra_index % 2 == 1

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = border
        if use_alt:
            cell.fill = alt_fill


def autofit_columns(ws: Worksheet, num_cols: int, *, max_width: int = 44) -> None:
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def finalize_sheet(ws: Worksheet, num_cols: int) -> None:
    autofit_columns(ws, num_cols)
