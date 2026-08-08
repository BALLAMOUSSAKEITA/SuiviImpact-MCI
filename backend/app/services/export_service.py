from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import extract, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.direction import Direction
from app.models.modules import Mission, Ppm, Projet, Recommandation
from app.models.plan_action import Activite, TachePlan
from app.models.tache import Tache, TacheStatut
from app.services.excel_branding import (
    finalize_sheet,
    prepare_branded_sheet,
    write_data_rows,
    write_row_cells,
)

TACHE_STATUT_LABEL: dict[TacheStatut, str] = {
    TacheStatut.EN_COURS: "En cours",
    TacheStatut.TERMINEE: "Terminée",
    TacheStatut.EN_RETARD: "En retard",
}


def _save_workbook(wb: Workbook) -> BytesIO:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _pao_avancement_pct(taches: list[Tache]) -> float:
    """Avancement = part pondérée des tâches terminées (aligné sur la logique suivi PAO)."""
    if not taches:
        return 0.0
    total = sum(float(t.ponderation) for t in taches)
    if total <= 0:
        return 0.0
    done = sum(
        float(t.ponderation) for t in taches if t.statut == TacheStatut.TERMINEE
    )
    return round(done / total * 100, 2)


async def _load_pao_export_rows(
    db: AsyncSession, annee: int
) -> tuple[list[tuple], list[tuple]]:
    """Retourne (lignes activités, lignes tâches) pour l'export PAO."""
    result = await db.execute(
        select(Activite)
        .where(
            Activite.date_debut.isnot(None),
            Activite.date_fin.isnot(None),
            extract("year", Activite.date_debut) == annee,
        )
        .options(
            selectinload(Activite.objectif),
            selectinload(Activite.directions),
        )
        .order_by(Activite.code)
    )
    activites = result.scalars().all()
    if not activites:
        return [], []

    activite_ids = [a.id for a in activites]
    taches_result = await db.execute(
        select(Tache)
        .where(
            Tache.activite_id.in_(activite_ids),
            Tache.tache_plan_id.isnot(None),
            Tache.annee == annee,
        )
        .order_by(Tache.activite_id, Tache.id)
    )
    taches_all = list(taches_result.scalars().all())
    taches_by_activite: dict[int, list[Tache]] = {}
    tache_plan_ids: set[int] = set()
    for t in taches_all:
        taches_by_activite.setdefault(t.activite_id, []).append(t)
        if t.tache_plan_id:
            tache_plan_ids.add(t.tache_plan_id)

    tp_map: dict[int, TachePlan] = {}
    if tache_plan_ids:
        tp_result = await db.execute(
            select(TachePlan).where(TachePlan.id.in_(tache_plan_ids))
        )
        tp_map = {tp.id: tp for tp in tp_result.scalars().all()}

    direction_ids = {d.direction_id for a in activites for d in a.directions}
    dir_map: dict[int, Direction] = {}
    if direction_ids:
        dir_result = await db.execute(
            select(Direction).where(Direction.id.in_(direction_ids))
        )
        dir_map = {d.id: d for d in dir_result.scalars().all()}

    activite_rows: list[tuple] = []
    tache_rows: list[tuple] = []

    for a in activites:
        if a.objectif is None:
            continue
        dir_id = a.directions[0].direction_id if a.directions else None
        direction = dir_map.get(dir_id) if dir_id else None
        if direction is None:
            continue

        taches = taches_by_activite.get(a.id, [])
        nb = len(taches)
        nb_terminees = sum(1 for t in taches if t.statut == TacheStatut.TERMINEE)
        avancement = _pao_avancement_pct(taches)

        activite_rows.append(
            (
                a.code,
                a.description,
                a.objectif.code,
                a.objectif.description,
                direction.code,
                direction.libelle,
                a.date_debut.isoformat() if a.date_debut else "",
                a.date_fin.isoformat() if a.date_fin else "",
                float(a.budget),
                nb,
                nb_terminees,
                avancement,
                float(a.execution),
                a.email_responsable or "",
                a.email_ministre or "",
            )
        )

        for t in taches:
            tp = tp_map.get(t.tache_plan_id) if t.tache_plan_id else None
            code_plan = tp.code if tp else ""
            desc_plan = tp.description if tp else t.description
            tache_rows.append(
                (
                    a.code,
                    a.description,
                    code_plan,
                    desc_plan,
                    t.description,
                    float(t.ponderation),
                    TACHE_STATUT_LABEL.get(t.statut, t.statut.value),
                    "Oui" if t.statut == TacheStatut.TERMINEE else "Non",
                    t.trimestre,
                    t.annee,
                    t.responsable,
                    t.observation or "",
                )
            )

    return activite_rows, tache_rows


async def export_pao(db: AsyncSession, annee: int) -> BytesIO:
    wb = Workbook()
    ws_act = wb.active
    ws_act.title = "Activités"
    headers_act = [
        "Code activité",
        "Description activité",
        "Code objectif",
        "Objectif",
        "Code direction",
        "Direction",
        "Date début",
        "Date fin",
        "Budget (GNF)",
        "Nb tâches",
        "Nb tâches terminées",
        "Avancement global (%)",
        "Exécution enregistrée (%)",
        "E-mail responsable",
        "E-mail ministre",
    ]
    start_act = prepare_branded_sheet(
        ws_act,
        report_title="Plan d'action opérationnel — Activités",
        subtitle=f"Année {annee} · SuiviImpact",
        headers=headers_act,
    )

    ws_taches = wb.create_sheet("Tâches")
    headers_taches = [
        "Code activité",
        "Description activité",
        "Code tâche plan",
        "Libellé tâche plan",
        "Description suivi",
        "Pondération (%)",
        "Statut suivi",
        "Réalisée",
        "Trimestre",
        "Année",
        "Responsable",
        "Observation",
    ]
    start_taches = prepare_branded_sheet(
        ws_taches,
        report_title="Plan d'action opérationnel — Tâches",
        subtitle=f"Année {annee} · SuiviImpact",
        headers=headers_taches,
    )

    activite_rows, tache_rows = await _load_pao_export_rows(db, annee)
    write_data_rows(ws_act, start_act, activite_rows)
    write_data_rows(ws_taches, start_taches, tache_rows)

    return _save_workbook(wb)


async def export_recommandations(db: AsyncSession) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommandations RCC"
    headers = [
        "Trimestre",
        "Année",
        "Date",
        "Description",
        "Responsable",
        "Exécution (%)",
        "Observations",
    ]
    start = prepare_branded_sheet(
        ws,
        report_title="Recommandations du Comité de coordination (RCC)",
        subtitle="Export SuiviImpact",
        headers=headers,
    )

    result = await db.execute(select(Recommandation).order_by(Recommandation.id))
    for i, item in enumerate(result.scalars().all()):
        write_row_cells(
            ws,
            start + i,
            (
                item.trimestre,
                item.annee,
                item.date_recommandation.isoformat(),
                item.description,
                item.responsable,
                float(item.execution),
                item.observations or "",
            ),
            zebra_index=i,
        )
    finalize_sheet(ws, len(headers))

    return _save_workbook(wb)


async def export_missions(db: AsyncSession) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Missions"
    headers = [
        "Trimestre",
        "Année",
        "Date",
        "Description",
        "Responsable",
        "Exécution (%)",
        "Observations",
    ]
    start = prepare_branded_sheet(
        ws,
        report_title="Suivi des missions",
        subtitle="Export SuiviImpact",
        headers=headers,
    )

    result = await db.execute(select(Mission).order_by(Mission.id))
    for i, item in enumerate(result.scalars().all()):
        write_row_cells(
            ws,
            start + i,
            (
                item.trimestre,
                item.annee,
                item.date_mission.isoformat(),
                item.description,
                item.responsable,
                float(item.execution),
                item.observations or "",
            ),
            zebra_index=i,
        )
    finalize_sheet(ws, len(headers))

    return _save_workbook(wb)


async def export_ppm(db: AsyncSession) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Marchés PPM"
    headers = [
        "Numéro",
        "Intitulé",
        "Type",
        "Mode passation",
        "Montant estimé",
        "Montant attribué",
        "Financement",
        "Date",
        "Statut",
    ]
    start = prepare_branded_sheet(
        ws,
        report_title="Plan de passation des marchés (PPM)",
        subtitle="Export SuiviImpact",
        headers=headers,
    )

    result = await db.execute(select(Ppm).order_by(Ppm.id))
    for i, item in enumerate(result.scalars().all()):
        write_row_cells(
            ws,
            start + i,
            (
                item.numero or "",
                item.intitule,
                item.type_marche or "",
                item.mode_passation or "",
                float(item.montant_estime or 0),
                float(item.montant_attribue or 0),
                item.financement or "",
                item.date_marche.isoformat() if item.date_marche else "",
                item.statut.value,
            ),
            zebra_index=i,
        )
    finalize_sheet(ws, len(headers))

    return _save_workbook(wb)


async def export_projets(db: AsyncSession) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projets"
    headers = [
        "Identifiant",
        "Nom du projet",
        "Abréviation",
        "Coût",
        "Bailleur",
        "Part État",
        "Part Bailleur",
        "Exécution financière (%)",
        "Exécution physique (%)",
        "Date début",
        "Date fin",
    ]
    start = prepare_branded_sheet(
        ws,
        report_title="Portefeuille projets",
        subtitle="Export SuiviImpact",
        headers=headers,
    )

    result = await db.execute(select(Projet).order_by(Projet.id))
    for i, item in enumerate(result.scalars().all()):
        write_row_cells(
            ws,
            start + i,
            (
                item.code,
                item.description,
                item.abreviation or "",
                float(item.cout or 0),
                item.bailleur or "",
                float(item.part_etat or 0),
                float(item.part_bailleur or 0),
                float(item.execution_financiere),
                float(item.execution_physique),
                item.date_debut.isoformat() if item.date_debut else "",
                item.date_fin.isoformat() if item.date_fin else "",
            ),
            zebra_index=i,
        )
    finalize_sheet(ws, len(headers))

    return _save_workbook(wb)
