from __future__ import annotations

import unicodedata
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.finance import FINANCE_SNAPSHOT_ID, FinanceLigne, FinanceSnapshot
from app.schemas.finance import FinanceLigneRead, FinanceSnapshotRead, FinanceStateRead

ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}
HEADER_MARKERS = ("titres du budget", "titre du budget")
SKIP_TITRES = {
    "titres du budget",
    "titre du budget",
    "montants en gnf",
    "taux de decaissement",
    "sources d information",
    "prevus / lfi (1)",
    "engages (2)",
    "payes (3)",
}


def _norm(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return " ".join(text.lower().replace("\xa0", " ").replace("'", " ").split())


def _parse_number(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = str(value).replace("\xa0", " ").strip()
    if not text or text in {"-", "—"}:
        return None
    text = text.replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _as_ratio(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    if value > Decimal("1.5"):
        return value / Decimal("100")
    return value


def _compute_ratio(numerateur: Decimal | None, denominateur: Decimal | None) -> Decimal | None:
    if numerateur is None or denominateur is None or denominateur == 0:
        return Decimal("0") if denominateur == 0 and numerateur is not None else None
    return numerateur / denominateur


def _is_total(titre: str) -> bool:
    n = _norm(titre)
    return n in {"total", "totaux"} or n.startswith("total ")


def _looks_like_header(titre: str) -> bool:
    n = _norm(titre)
    if n in SKIP_TITRES:
        return True
    if "titres du budget" in n or "montants en gnf" in n:
        return True
    if "prevus" in n and "lfi" in n:
        return True
    return False


def parse_finance_workbook(content: bytes) -> list[dict]:
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Fichier Excel illisible",
        ) from exc

    sheet = workbook.active
    if sheet is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le fichier Excel ne contient aucune feuille",
        )

    max_row = sheet.max_row or 1
    max_col = min(sheet.max_column or 1, 16)

    header_row = 1
    found = False
    for row_idx in range(1, min(16, max_row) + 1):
        for col_idx in range(1, max_col + 1):
            cell_norm = _norm(sheet.cell(row_idx, col_idx).value or "")
            if any(marker in cell_norm for marker in HEADER_MARKERS):
                header_row = row_idx
                found = True
                break
        if found:
            break

    cols = {
        "titre": 1,
        "prevu": 2,
        "engage": 3,
        "paye": 4,
        "taux_eng": 5,
        "taux_caisse": 6,
        "source": 7,
    }
    for row_idx in (header_row, header_row + 1):
        if row_idx > max_row:
            continue
        for col_idx in range(1, max_col + 1):
            n = _norm(sheet.cell(row_idx, col_idx).value or "")
            if not n:
                continue
            if any(marker in n for marker in HEADER_MARKERS) or n == "titre":
                cols["titre"] = col_idx
            elif "prevu" in n or "lfi" in n:
                cols["prevu"] = col_idx
            elif "engag" in n and "base" not in n and "%" not in n and "(2) / (1)" not in n.replace(" ", ""):
                cols["engage"] = col_idx
            elif ("paye" in n or "payes" in n) and "base" not in n and "%" not in n:
                cols["paye"] = col_idx
            elif "caisse" in n:
                cols["taux_caisse"] = col_idx
            elif "engagement" in n:
                cols["taux_eng"] = col_idx
            elif "source" in n:
                cols["source"] = col_idx

    lignes: list[dict] = []
    ordre = 0
    for row_idx in range(header_row + 1, max_row + 1):
        titre_raw = sheet.cell(row_idx, cols["titre"]).value
        if titre_raw is None:
            continue
        titre = str(titre_raw).replace("\xa0", " ").strip()
        if not titre or _looks_like_header(titre):
            continue

        prevu = _parse_number(sheet.cell(row_idx, cols["prevu"]).value)
        engage = _parse_number(sheet.cell(row_idx, cols["engage"]).value)
        paye = _parse_number(sheet.cell(row_idx, cols["paye"]).value)
        taux_eng = _as_ratio(_parse_number(sheet.cell(row_idx, cols["taux_eng"]).value))
        taux_caisse = _as_ratio(_parse_number(sheet.cell(row_idx, cols["taux_caisse"]).value))
        source_raw = sheet.cell(row_idx, cols["source"]).value
        source = str(source_raw).replace("\xa0", " ").strip() if source_raw is not None else None
        if source == "":
            source = None

        is_total = _is_total(titre)
        if (
            not is_total
            and prevu is None
            and engage is None
            and paye is None
            and taux_eng is None
            and taux_caisse is None
        ):
            continue

        if taux_eng is None:
            taux_eng = _compute_ratio(engage, prevu)
        if taux_caisse is None:
            taux_caisse = _compute_ratio(paye, prevu)

        ordre += 1
        lignes.append(
            {
                "ordre": ordre,
                "titre_budget": titre,
                "montant_prevu": prevu,
                "montant_engage": engage,
                "montant_paye": paye,
                "taux_engagement": taux_eng,
                "taux_caisse": taux_caisse,
                "source_information": source,
                "is_total": is_total,
            }
        )

    if not lignes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Aucune ligne budgétaire trouvée dans le fichier",
        )
    return lignes


async def get_state(db: AsyncSession) -> FinanceStateRead:
    snapshot_result = await db.execute(
        select(FinanceSnapshot).where(FinanceSnapshot.id == FINANCE_SNAPSHOT_ID)
    )
    snapshot = snapshot_result.scalar_one_or_none()
    lignes_result = await db.execute(
        select(FinanceLigne).order_by(FinanceLigne.ordre, FinanceLigne.id)
    )
    lignes = list(lignes_result.scalars().all())
    return FinanceStateRead(
        snapshot=FinanceSnapshotRead.model_validate(snapshot) if snapshot else None,
        lignes=[FinanceLigneRead.model_validate(row) for row in lignes],
    )


async def import_excel(db: AsyncSession, file: UploadFile, user_id: int) -> FinanceStateRead:
    filename = file.filename or "import.xlsx"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Fichier Excel (.xlsx) requis",
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Fichier vide",
        )

    parsed = parse_finance_workbook(content)

    await db.execute(delete(FinanceLigne))

    snapshot_result = await db.execute(
        select(FinanceSnapshot).where(FinanceSnapshot.id == FINANCE_SNAPSHOT_ID)
    )
    snapshot = snapshot_result.scalar_one_or_none()
    now = datetime.now(UTC)
    if snapshot is None:
        snapshot = FinanceSnapshot(
            id=FINANCE_SNAPSHOT_ID,
            filename=filename[:255],
            imported_at=now,
            imported_by_user_id=user_id,
            row_count=len(parsed),
        )
        db.add(snapshot)
    else:
        snapshot.filename = filename[:255]
        snapshot.imported_at = now
        snapshot.imported_by_user_id = user_id
        snapshot.row_count = len(parsed)

    for row in parsed:
        db.add(FinanceLigne(**row))

    await db.commit()
    return await get_state(db)
