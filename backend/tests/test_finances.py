from io import BytesIO

import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.user import AccessType, User, UserRole


def _build_budget_xlsx(
    rows: list[tuple],
    *,
    with_header: bool = True,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    if with_header:
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:D1")
        ws.merge_cells("E1:F1")
        ws.merge_cells("G1:G2")
        ws["A1"] = "Titres du budget"
        ws["B1"] = "Montants en GNF"
        ws["E1"] = "Taux de décaissement"
        ws["G1"] = "Sources d'information"
        ws["B2"] = "Prévus / LFI (1)"
        ws["C2"] = "Engagés (2)"
        ws["D2"] = "Payés (3)"
        ws["E2"] = 'Base "engagement" (2) /(1) en %'
        ws["F2"] = 'Base "caisse" (3)/(1) en %'
        start = 3
    else:
        start = 1
    for offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            ws.cell(start + offset, col, value)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


SAMPLE_ROWS = [
    (
        "Titre II : Dépense du personnel",
        63265378755,
        6384925709,
        5639035022,
        0.10092290340544884,
        0.08913303188838231,
        "Chaine de Dépenses",
    ),
    (
        "Titre III : Dépense de Biens et Service",
        67051009296,
        3098079072,
        0,
        0.04620480891381331,
        0,
        "Chaine de Dépenses",
    ),
    (
        "Titre IV : Dépense de Transfert",
        24568747329,
        2752325000,
        0,
        0.11202545099852372,
        0,
        "Chaine de Dépenses",
    ),
    (
        "Titre V : Dépense d'Investissement",
        479360885536,
        9313017088,
        0,
        0.019427987074052983,
        0,
        "Chaine de Dépenses",
    ),
    ("Total", 634246020916, 21548346869, 5639035022, 0.03397474506482379, 0.00889092692115894, None),
]


@pytest.mark.asyncio
async def test_get_finances_requires_auth(client: AsyncClient):
    response = await client.get("/api/v1/finances")
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_get_finances_empty(client: AsyncClient, auth_headers: dict[str, str]):
    response = await client.get("/api/v1/finances", headers=auth_headers)
    assert response.status_code == 200
    body = response.json()
    assert body["snapshot"] is None
    assert body["lignes"] == []


@pytest.mark.asyncio
async def test_import_xlsx_and_overwrite(client: AsyncClient, auth_headers: dict[str, str]):
    first = await client.post(
        "/api/v1/finances/import",
        headers=auth_headers,
        files={
            "file": (
                "budget.xlsx",
                _build_budget_xlsx(SAMPLE_ROWS),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert first.status_code == 200, first.text
    body = first.json()
    assert body["snapshot"]["filename"] == "budget.xlsx"
    assert body["snapshot"]["row_count"] == 5
    assert len(body["lignes"]) == 5
    assert body["lignes"][0]["titre_budget"].startswith("Titre II")
    assert body["lignes"][-1]["is_total"] is True
    assert float(body["lignes"][0]["montant_prevu"]) == 63265378755
    assert abs(float(body["lignes"][0]["taux_engagement"]) - 0.10092290340544884) < 1e-6

    second_rows = [
        ("Titre II : Personnel", 1000, 200, 100, None, None, "Chaîne"),
        ("Total", 1000, 200, 100, None, None, None),
    ]
    second = await client.post(
        "/api/v1/finances/import",
        headers=auth_headers,
        files={
            "file": (
                "budget2.xlsx",
                _build_budget_xlsx(second_rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert second.status_code == 200, second.text
    overwritten = second.json()
    assert overwritten["snapshot"]["filename"] == "budget2.xlsx"
    assert len(overwritten["lignes"]) == 2
    assert float(overwritten["lignes"][0]["montant_prevu"]) == 1000
    assert abs(float(overwritten["lignes"][0]["taux_engagement"]) - 0.2) < 1e-6
    assert abs(float(overwritten["lignes"][0]["taux_caisse"]) - 0.1) < 1e-6


@pytest.mark.asyncio
async def test_import_rejects_invalid_file(client: AsyncClient, auth_headers: dict[str, str]):
    response = await client.post(
        "/api/v1/finances/import",
        headers=auth_headers,
        files={"file": ("notes.txt", b"pas un excel", "text/plain")},
    )
    assert response.status_code == 400


@pytest.mark.asyncio
async def test_import_requires_write(
    client: AsyncClient, db_session: AsyncSession
):
    reader = User(
        username="lecteur_finances",
        password_hash=hash_password("lecteur123"),
        prenom="Lecteur",
        nom="",
        role=UserRole.USER,
        type_acces=AccessType.LECTURE,
        etat=True,
    )
    db_session.add(reader)
    await db_session.commit()

    login = await client.post(
        "/api/v1/auth/login",
        json={"username": "lecteur_finances", "password": "lecteur123"},
    )
    assert login.status_code == 200
    headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

    response = await client.post(
        "/api/v1/finances/import",
        headers=headers,
        files={
            "file": (
                "budget.xlsx",
                _build_budget_xlsx(SAMPLE_ROWS),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 403
