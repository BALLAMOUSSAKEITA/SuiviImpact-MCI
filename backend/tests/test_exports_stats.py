import pytest
from httpx import AsyncClient


EXPORT_PATHS = [
    "/api/v1/exports/pao?mode=annee&annee=2025",
    "/api/v1/exports/recommandations",
    "/api/v1/exports/missions",
    "/api/v1/exports/ppm",
    "/api/v1/exports/projets",
]

STATS_PATHS = [
    "/api/v1/stats/activites",
    "/api/v1/stats/recommandations",
    "/api/v1/stats/missions",
    "/api/v1/stats/ppm",
    "/api/v1/stats/projets",
]


@pytest.mark.asyncio
async def test_export_pao_with_activite(client: AsyncClient, auth_headers: dict[str, str]):
    from tests.helpers import (
        create_direction,
        create_objectif,
        create_tache_plan,
        pao_payload,
        post_pao,
    )

    direction_id = await create_direction(client, auth_headers)
    objectif_id = await create_objectif(client, auth_headers, "OC-EXP")
    tache_plan_id = await create_tache_plan(client, auth_headers, "TP-EXP")
    create = await post_pao(
        client,
        auth_headers,
        pao_payload(
            objectif_id=objectif_id,
            direction_id=direction_id,
            tache_plan_id=tache_plan_id,
        ),
    )
    assert create.status_code == 201

    response = await client.get(
        "/api/v1/exports/pao?mode=annee&annee=2025",
        headers=auth_headers,
    )
    assert response.status_code == 200
    assert "spreadsheet" in response.headers.get("content-type", "")
    assert len(response.content) > 200

    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.excel_branding import EXPORT_MINISTRY

    wb = load_workbook(BytesIO(response.content))
    assert wb["Activités"]["A1"].value == EXPORT_MINISTRY
    assert "Plan d" in str(wb["Activités"]["A3"].value or "")

    # Filtre mois : date_debut 2025-01-15 → inclus pour janvier 2025 uniquement
    jan = await client.get(
        "/api/v1/exports/pao?mode=mois&mois=2025-01",
        headers=auth_headers,
    )
    assert jan.status_code == 200
    wb_jan = load_workbook(BytesIO(jan.content))
    assert wb_jan["Activités"].max_row >= 6  # en-tête + au moins une ligne

    fev = await client.get(
        "/api/v1/exports/pao?mode=mois&mois=2025-02",
        headers=auth_headers,
    )
    assert fev.status_code == 200
    wb_fev = load_workbook(BytesIO(fev.content))
    # Pas d'activité avec date_debut en février
    assert wb_fev["Activités"].max_row == 5  # en-tête branding seulement


@pytest.mark.asyncio
@pytest.mark.parametrize("path", EXPORT_PATHS)
async def test_exports_return_excel(client: AsyncClient, auth_headers: dict[str, str], path: str):
    response = await client.get(path, headers=auth_headers)
    assert response.status_code == 200
    assert "spreadsheet" in response.headers.get("content-type", "")
    assert len(response.content) > 100


@pytest.mark.asyncio
@pytest.mark.parametrize("path", STATS_PATHS)
async def test_stats_endpoints(client: AsyncClient, auth_headers: dict[str, str], path: str):
    response = await client.get(path, headers=auth_headers)
    assert response.status_code == 200
    assert isinstance(response.json(), dict)
